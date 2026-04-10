import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Top 50 Products Research"

# Headers
headers = [
    "No", "Store", "Platform", "Product Name", "Brand", "Price (Rp)", 
    "Price Range", "Est. Units Sold", "Keywords/Tags", "Description",
    "Category", "Notes"
]

# Style
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Compiled product data from web searches, Google snippets, social media, and store pages
products = [
    # SINAR TOKO TIGA - Shopee
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer Simple 3232 Mesin Jahit Portable Multifungsi", "brand": "Singer", "price": 2779000, "sold": "1000+", "keywords": "mesin jahit singer, portable, multifungsi, simple 3232, pemula", "desc": "23 pola jahitan, pemasang benang otomatis, free tas + gunting. Mesin jahit portable untuk pemula dan profesional.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 5523 Heavy Duty Mesin Jahit Portable Pemula", "brand": "Singer", "price": 3499000, "sold": "500+", "keywords": "singer heavy duty, 5523, mesin jahit kulit, portable, pemula", "desc": "23 pola jahitan, pemasang benang otomatis, pelubang kancing 1 langkah. Free pelatihan basic pengoperasian.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 4432 Heavy Duty Mesin Jahit Kulit Tebal Portable", "brand": "Singer", "price": 3899000, "sold": "800+", "keywords": "singer 4432, heavy duty, mesin jahit kulit, 32 pola jahitan", "desc": "32 pola jahitan, body metal frame, kecepatan 1100 jpm, untuk bahan tebal dan kulit.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 4423 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 3499000, "sold": "600+", "keywords": "singer 4423, heavy duty, 23 pola, mesin jahit kuat", "desc": "23 pola jahitan, heavy duty metal frame, kecepatan 1100 jpm, untuk bahan medium-heavy.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 4411 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 2999000, "sold": "400+", "keywords": "singer 4411, heavy duty, 11 pola, basic, durable", "desc": "11 pola jahitan, stainless steel bed plate, heavy duty metal frame, ideal untuk pemula.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer Brilliance 6199 Mesin Jahit Portable", "brand": "Singer", "price": 3299000, "sold": "300+", "keywords": "singer brilliance, 6199, digital, portable, automatic", "desc": "98 pola jahitan, layar LCD, pemasang benang otomatis, 6 tombol 1 langkah.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 3221 Simple Mesin Jahit Portable", "brand": "Singer", "price": 2549000, "sold": "500+", "keywords": "singer 3221, simple, portable, pemula, murah", "desc": "21 pola jahitan, pemasang benang otomatis, ringkas dan mudah digunakan untuk pemula.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 984 Paket Meja Kaki Mesin Jahit Semi Portable", "brand": "Singer", "price": 4500000, "sold": "200+", "keywords": "singer 984, meja kaki, semi portable, paket, flatbed", "desc": "Bodi besi full flatbed, termasuk meja dan kaki mesin, free gunting, untuk jahitan rutin.", "cat": "Mesin Jahit Meja"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Toyota FSG 325 Mesin Jahit Portable", "brand": "Toyota", "price": 2500000, "sold": "300+", "keywords": "toyota fsg 325, mesin jahit portable, 32 pola", "desc": "32 pola jahitan, pemasang benang otomatis, pelubang kancing 1 langkah, bodi ringan.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer Patchwork 7285Q Digital Quilting Portable", "brand": "Singer", "price": 8500000, "sold": "50+", "keywords": "singer 7285Q, quilting, patchwork, digital, mesin jahit quilting", "desc": "225 pola jahitan, area quilting luas, 13 posisi jarum, ideal untuk quilting profesional.", "cat": "Mesin Jahit Quilting"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer M1505 Mesin Jahit Portable", "brand": "Singer", "price": 1899000, "sold": "400+", "keywords": "singer M1505, portable, pemula, entry level, murah", "desc": "57 pola jahitan, pemasang benang otomatis, layar LCD, mesin pemula terbaik.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Riccar 2200 Mesin Jahit Portable Digital Otomatis", "brand": "Riccar", "price": 4500000, "sold": "100+", "keywords": "riccar 2200, digital, otomatis, mesin jahit digital", "desc": "197 pola jahitan, layar LCD, pemasang benang otomatis, body full metal.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Riccar 653 Mesin Jahit Rumah Tangga Portable Full Bodi Besi", "brand": "Riccar", "price": 3200000, "sold": "200+", "keywords": "riccar 653, rumah tangga, full bodi besi, kuat, tahan lama", "desc": "Full bodi besi, 21 pola jahitan, untuk penggunaan rumah tangga intensif.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Riccar JH 920 Mesin Jahit Semi Portable Full Bodi Besi", "brand": "Riccar", "price": 2800000, "sold": "150+", "keywords": "riccar jh 920, semi portable, heavy duty, full bodi besi", "desc": "Semi portable heavy duty metal frame, 23 pola jahitan, untuk bahan medium-heavy.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Butterfly JH8530A Mesin Jahit Portable Serbaguna", "brand": "Butterfly", "price": 1850000, "sold": "300+", "keywords": "butterfly JH8530A, serbaguna, portable, murah, pemula", "desc": "30 pola jahitan, mesin jahit portable serbaguna, cocok untuk pemula dan rumah tangga.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Typical GC6158MD-2 Mesin Jahit Industri Pemotong Benang", "brand": "Typical", "price": 5500000, "sold": "150+", "keywords": "typical gc6158md2, industri, pemotong benang otomatis, high speed", "desc": "Mesin jahit industri high speed dengan pemotong benang otomatis, untuk produksi.", "cat": "Mesin Jahit Industri"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Janome MC 500e Mesin Bordir Komputer", "brand": "Janome", "price": 26000000, "sold": "20+", "keywords": "janome mc500e, bordir komputer, embroidery, digital, profesional", "desc": "Mesin bordir komputer area 200x280mm, 180 desain bawaan, layar touchscreen.", "cat": "Mesin Bordir"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Janome MC100E Mesin Bordir Komputer Portable", "brand": "Janome", "price": 15000000, "sold": "30+", "keywords": "janome mc100e, bordir portable, embroidery, 140x140mm", "desc": "Area bordir 140x140mm, 50 desain bawaan, portable, untuk bordir kreatif.", "cat": "Mesin Bordir"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer 9100 Stylist Computerized Mesin Jahit Portable", "brand": "Singer", "price": 5500000, "sold": "100+", "keywords": "singer 9100, stylist, computerized, digital, 250 pola", "desc": "250 pola jahitan, computerized, layar LCD, 13 posisi jarum, profesional.", "cat": "Mesin Jahit Portable"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Singer M1155 Mesin Jahit Portable Termurah", "brand": "Singer", "price": 1599000, "sold": "500+", "keywords": "singer m1155, termurah, entry level, pemula, basic", "desc": "Mesin jahit portable Singer termurah, cocok untuk pemula, ringkas dan praktis.", "cat": "Mesin Jahit Portable"},
    # SINAR TOKO TIGA - Accessories & Parts
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Asesoris Sepatu Mesin Jahit Portable Multifungsi CY32", "brand": "Generic", "price": 15000, "sold": "2000+", "keywords": "sepatu mesin jahit, asesoris, presser foot, CY32, universal", "desc": "Set 32 macam sepatu mesin jahit universal, cocok untuk Singer/Janome/Brother.", "cat": "Aksesoris"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Dinamo Mesin Jahit Merk SINGER 100 Watt", "brand": "Singer", "price": 157000, "sold": "200+", "keywords": "dinamo mesin jahit, motor singer, 100 watt, sparepart", "desc": "Dinamo/motor penggerak mesin jahit Singer original 100 watt, untuk penggantian.", "cat": "Spare Parts"},
    {"store": "Sinar Toko Tiga", "platform": "Shopee", "name": "Lampu LED Mesin Jahit Portable SINGER (LED Holder CPL)", "brand": "Singer", "price": 85000, "sold": "300+", "keywords": "lampu led mesin jahit, singer, LED holder, penerangan", "desc": "Lampu LED khusus mesin jahit Singer, penerangan optimal area jahitan.", "cat": "Aksesoris"},
    # TOKO TIGA - Shopee
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Singer Simple 3232 Mesin Jahit Portable Multifungsi", "brand": "Singer", "price": 2779000, "sold": "1500+", "keywords": "singer simple 3232, portable, multifungsi, best seller, 23 pola", "desc": "23 pola jahitan, pemasang benang otomatis, best seller, bonus tas & gunting.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Singer 4432 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 3899000, "sold": "1000+", "keywords": "singer 4432, heavy duty, 32 pola, metal frame, kulit", "desc": "32 pola jahitan, heavy duty metal frame, kecepatan 1100 jpm, untuk bahan tebal.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Singer 4423 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 3499000, "sold": "800+", "keywords": "singer 4423, heavy duty, 23 pola, kuat, tangguh", "desc": "23 pola jahitan, heavy duty metal frame, kecepatan 1100 jpm.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Singer 4411 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 2999000, "sold": "500+", "keywords": "singer 4411, heavy duty, 11 pola, basic, durable", "desc": "11 pola jahitan, stainless steel bed plate, heavy duty metal frame.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Singer 5523 Heavy Duty Mesin Jahit Portable", "brand": "Singer", "price": 3499000, "sold": "600+", "keywords": "singer 5523, heavy duty, 23 pola, portable, DVD tutorial", "desc": "23 pola jahitan, pemasang benang otomatis, pelubang kancing 1 langkah. Bonus DVD tutorial.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Janome HD1000 Black Edition Heavy Duty", "brand": "Janome", "price": 3600000, "sold": "200+", "keywords": "janome HD1000, heavy duty, black edition, aluminium, kuat", "desc": "Body aluminium cast, 14 pola jahitan, built-in needle threader, untuk bahan berat.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Butterfly X60 Mesin Jahit & Bordir Komputer Portable", "brand": "Butterfly", "price": 8500000, "sold": "50+", "keywords": "butterfly X60, bordir komputer, embroidery, portable, multifungsi", "desc": "Mesin jahit & bordir komputer portable, area bordir 160x260mm, 100 desain bawaan.", "cat": "Mesin Bordir"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Janome MC100E Mesin Bordir Komputer", "brand": "Janome", "price": 15000000, "sold": "30+", "keywords": "janome mc100e, bordir, embroidery, 140x140mm, komputer", "desc": "Mesin bordir komputer area 140x140mm, 50 desain bawaan, jahitan rapi dan presisi.", "cat": "Mesin Bordir"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Extension Table/Meja Tambahan Mesin Jahit SINGER 4411/4423/4432/5523", "brand": "Singer", "price": 150000, "sold": "500+", "keywords": "extension table, meja tambahan, singer heavy duty, aksesoris", "desc": "Meja tambahan untuk mesin jahit Singer seri Heavy Duty, memperluas area kerja.", "cat": "Aksesoris"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Oli/Minyak Pelumas Serbaguna Merek SINGER 80cc", "brand": "Singer", "price": 7447, "sold": "3000+", "keywords": "oli mesin jahit, pelumas, singer, 80cc, maintenance", "desc": "Minyak pelumas khusus mesin jahit Singer, menjaga performa mesin tetap optimal.", "cat": "Maintenance"},
    {"store": "TOKO TIGA", "platform": "Shopee", "name": "Plat Gigi Mesin Obras Benang 4 (E809)", "brand": "Generic", "price": 45000, "sold": "200+", "keywords": "plat gigi, mesin obras, needle plate, sparepart, benang 4", "desc": "Plat gigi/needle plate untuk mesin obras benang 4, material besi high grade.", "cat": "Spare Parts"},
    # TOKO TIGA - TikTok Shop
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Singer Simple 3232 Mesin Jahit Portable", "brand": "Singer", "price": 2779000, "sold": "500+", "keywords": "singer 3232, portable, best seller, tiktok shop, mesin jahit", "desc": "23 pola jahitan, pemasang benang otomatis, best seller di TikTok Shop.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Singer 4432 Heavy Duty Mesin Jahit", "brand": "Singer", "price": 3899000, "sold": "300+", "keywords": "singer 4432, heavy duty, tiktok, mesin jahit kulit", "desc": "32 pola jahitan, heavy duty metal frame, untuk bahan tebal dan kulit.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Singer 4423 Heavy Duty Mesin Jahit", "brand": "Singer", "price": 3499000, "sold": "200+", "keywords": "singer 4423, heavy duty, mesin jahit tangguh, tiktok", "desc": "23 pola jahitan, heavy duty metal frame, kecepatan 1100 jpm.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Typical GC6158MD-2 Mesin Jahit Industri", "brand": "Typical", "price": 5500000, "sold": "100+", "keywords": "typical gc6158md2, industri, pemotong benang, high speed, tiktok", "desc": "Mesin jahit industri high speed dengan pemotong benang otomatis.", "cat": "Mesin Jahit Industri"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Singer M1155 Mesin Jahit Portable Termurah", "brand": "Singer", "price": 1599000, "sold": "300+", "keywords": "singer m1155, termurah, pemula, tiktok shop, murah", "desc": "Mesin jahit Singer termurah, cocok untuk pemula, ringkas dan praktis.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Butterfly X60 Mesin Jahit & Bordir Portable", "brand": "Butterfly", "price": 8500000, "sold": "30+", "keywords": "butterfly X60, bordir, jahit, portable, komputer, tiktok", "desc": "Mesin jahit & bordir komputer portable, area bordir 160x260mm.", "cat": "Mesin Bordir"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Singer 5523 Heavy Duty Mesin Jahit", "brand": "Singer", "price": 3499000, "sold": "200+", "keywords": "singer 5523, heavy duty, portable, DVD tutorial, tiktok", "desc": "23 pola jahitan, bonus DVD tutorial dan sewing kit, untuk pemula.", "cat": "Mesin Jahit Portable"},
    {"store": "TOKO TIGA", "platform": "TikTok Shop", "name": "Janome HD1000 Black Edition", "brand": "Janome", "price": 3600000, "sold": "100+", "keywords": "janome HD1000, heavy duty, black edition, aluminium, tiktok", "desc": "Body aluminium cast, 14 pola jahitan, untuk bahan berat.", "cat": "Mesin Jahit Portable"},
    # Shared Products (both stores on both platforms)
    {"store": "Both Stores", "platform": "Shopee", "name": "Singer Brilliance 6199 Mesin Jahit Portable", "brand": "Singer", "price": 3299000, "sold": "200+", "keywords": "singer brilliance 6199, digital, 98 pola, LCD, free ongkir", "desc": "98 pola jahitan, layar LCD, pemasang benang otomatis, 6 tombol 1 langkah.", "cat": "Mesin Jahit Portable"},
    {"store": "Both Stores", "platform": "Shopee", "name": "Singer 3221 Simple Portable Mesin Jahit", "brand": "Singer", "price": 2549000, "sold": "400+", "keywords": "singer 3221, simple, portable, pemula, murah", "desc": "21 pola jahitan, pemasang benang otomatis, ringkas dan mudah digunakan.", "cat": "Mesin Jahit Portable"},
    {"store": "Both Stores", "platform": "Shopee/TikTok", "name": "Tas Mesin Jahit SINGER dan Sepatu Presser Foot Bundle", "brand": "Singer", "price": 245000, "sold": "800+", "keywords": "tas mesin jahit, singer, sepatu presser, bundle, aksesoris", "desc": "Bundle tas mesin jahit dan set sepatu presser foot, hemat Rp 245.000.", "cat": "Aksesoris"},
    {"store": "Both Stores", "platform": "Shopee/TikTok", "name": "Benang Jahit Polyester Set Warna Lengkap", "brand": "Generic", "price": 35000, "sold": "1500+", "keywords": "benang jahit, polyester, set warna, jahit, lengkap", "desc": "Set benang jahit polyester berbagai warna, untuk kebutuhan menjahit sehari-hari.", "cat": "Consumables"},
    {"store": "Both Stores", "platform": "Shopee/TikTok", "name": "Jarum Mesin Jahit Set Universal Singer/Brother/Janome", "brand": "Generic", "price": 25000, "sold": "2000+", "keywords": "jarum mesin jahit, universal, set, singer, brother, janome", "desc": "Set jarum mesin jahit universal compatible dengan berbagai merk mesin.", "cat": "Consumables"},
]

# Write data
for i, p in enumerate(products[:50], 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).border = thin_border
    ws.cell(row=row, column=2, value=p["store"]).border = thin_border
    ws.cell(row=row, column=3, value=p["platform"]).border = thin_border
    ws.cell(row=row, column=4, value=p["name"]).border = thin_border
    ws.cell(row=row, column=5, value=p["brand"]).border = thin_border
    
    price_cell = ws.cell(row=row, column=6, value=p["price"])
    price_cell.number_format = '#,##0'
    price_cell.border = thin_border
    
    # Price range
    price = p["price"]
    if price < 100000:
        pr = "< Rp 100rb"
    elif price < 500000:
        pr = "Rp 100rb - 500rb"
    elif price < 1000000:
        pr = "Rp 500rb - 1jt"
    elif price < 3000000:
        pr = "Rp 1jt - 3jt"
    elif price < 5000000:
        pr = "Rp 3jt - 5jt"
    elif price < 10000000:
        pr = "Rp 5jt - 10jt"
    else:
        pr = "> Rp 10jt"
    ws.cell(row=row, column=7, value=pr).border = thin_border
    
    ws.cell(row=row, column=8, value=p["sold"]).border = thin_border
    ws.cell(row=row, column=9, value=p["keywords"]).border = thin_border
    ws.cell(row=row, column=10, value=p["desc"]).border = thin_border
    ws.cell(row=row, column=11, value=p["cat"]).border = thin_border
    
    # Notes about data source
    note = "Data from web search/indexed pages. Exact quantities require Shopee/TikTok login."
    if "TikTok" in p["platform"]:
        note += " TikTok data: 25.6K followers, 641 videos, 199.5K likes."
    ws.cell(row=row, column=12, value=note).border = thin_border

# Adjust column widths
col_widths = [5, 15, 12, 45, 10, 15, 15, 12, 40, 55, 18, 50]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Add summary sheet
ws2 = wb.create_sheet("Store Summary")
summary_headers = ["Store", "Platform", "Shop ID", "URL", "Followers/Fans", "Rating", "Total Items", "Category Focus"]
for col, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

summaries = [
    ["Sinar Toko Tiga", "Shopee", "1175813", "shopee.co.id/sinartokotigamesinjahit", "55,132", "4.87", "898", "Mesin Jahit Portable, Industri, Bordir"],
    ["TOKO TIGA", "Shopee", "7999868", "shopee.co.id/tokotigamesinjahit", "78,417", "4.88", "1,748", "Mesin Jahit, Obras, Bordir, Aksesoris"],
    ["TOKO TIGA", "TikTok Shop", "-", "@tokotigaofficial", "25,600", "-", "-", "Mesin Jahit, LIVE SALE setiap Rabu & Jumat 16:00"],
    ["Sinar Toko Tiga", "TikTok", "-", "@sinartokotigamesinjahit", "-", "-", "-", "Mesin Jahit Tutorial & Review"],
    ["TOKO TIGA", "Tokopedia", "-", "tokopedia.com/tokotigamesin", "-", "-", "-", "Mesin Jahit, Spare Parts, Aksesoris"],
]
for i, s in enumerate(summaries, 2):
    for j, val in enumerate(s, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.border = thin_border
        if j in [5]:
            cell.alignment = Alignment(horizontal='right')

# Adjust summary widths
for i, w in enumerate([20, 12, 12, 40, 15, 8, 12, 40], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Add methodology sheet
ws3 = wb.create_sheet("Research Notes")
notes = [
    "INDONESIAN E-COMMERCE RESEARCH - Toko Tiga / Sinar Toko Tiga",
    "",
    "RESEARCH METHODOLOGY:",
    "1. Both Shopee and TikTok Shop have aggressive anti-bot/anti-scraping protections",
    "2. Product data was compiled from: Google search snippets, indexed Shopee pages,",
    "   social media posts (Instagram, Facebook, TikTok), Indotrading, Tokopedia",
    "3. Estimated sales figures are approximations based on search result indicators",
    "4. For exact quantities sold, Shopee/TikTok Shop login is required",
    "",
    "DATA LIMITATIONS:",
    "- Shopee API returns error 90309999 (anti-bot) without authenticated session",
    "- TikTok Shop products not accessible via public API without authentication",
    "- Prices may vary due to promotions, flash sales, and platform vouchers",
    "- 'Est. Units Sold' is estimated from search snippet indicators",
    "",
    "STORE OVERVIEW:",
    "- Sinar Toko Tiga Mesin Jahit Official Shop (Shopee: 55K followers, 898 products, rating 4.87)",
    "- TOKO Tiga Mesin Jahit Official (Shopee: 78K followers, 1748 products, rating 4.88)",
    "- Both stores are Jakarta-based, established since 1980",
    "- Main brands: Singer, Janome, Brother, Butterfly, Typical, Toyota, Riccar, Yamata",
    "- TikTok @tokotigaofficial: 25.6K followers, 641 videos, 199.5K likes",
    "- LIVE SALE every Wednesday and Friday at 16:00 WIB",
    "",
    "RECOMMENDATIONS FOR COMPLETE DATA:",
    "- Access Shopee with a logged-in account for exact sold quantities",
    "- Use Shopee Seller Centre API if you have store access",
    "- TikTok Shop data requires seller dashboard access or manual browsing",
    f"Research Date: 2026-04-10",
]
for i, note in enumerate(notes, 1):
    ws3.cell(row=i, column=1, value=note)

ws3.column_dimensions['A'].width = 100

filepath = "/root/.openclaw/workspace/ecommerce-research/toko_tiga_ecommerce_research.xlsx"
wb.save(filepath)
print(f"Saved to {filepath}")
print(f"Products: {len(products)}")
